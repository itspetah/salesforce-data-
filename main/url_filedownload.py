import os
import requests
from openpyxl import load_workbook
from requests.exceptions import ConnectionError, HTTPError

# Constants
DOWNLOAD_FOLDER = 'attachments/'  # Ensure this folder exists

# Load the Excel workbook and sheet
wb = load_workbook('File_URLs.xlsx')
ws = wb.active

# Access token
headers = {
    'Authorization': 'Bearer <>'
}

# Process each row in the Excel sheet
for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
    file_name, file_url = row
    file_path = os.path.join(DOWNLOAD_FOLDER, file_name)

    try:
        response = requests.get(file_url, headers=headers)
        response.raise_for_status()

        # Save the image to the specified path
        with open(file_path, 'wb') as f:
            f.write(response.content)
        print(f"Downloaded image: {file_name}")

    except ConnectionError as e:
        print(f"Connection error: {e}")
    except HTTPError as e:
        print(f"HTTP error: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
